#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Poliport Kontrol Formu Arşivi v1.0
-----------------------------------
Araç Kayıt Sorgulama raporundaki her araç için SSRS kontrol formunu
otomatik indirir, isimlendirir ve tarih bazlı klasöre kaydeder.

Klasör yapısı:
  Kontrol Formları/
    2026/
      Nisan/
        29/
          34ES2792_34GTL115_SEDAT_GUN.xlsx
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import re
import urllib.parse
from datetime import datetime
import json
import warnings

warnings.filterwarnings("ignore")

# ── Bağımlılık kontrolü ──────────────────────────────────────────────────────
def check_and_install(pkg, import_name=None):
    import_name = import_name or pkg
    try:
        __import__(import_name)
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg,
                               "--break-system-packages", "-q"])

check_and_install("openpyxl")
check_and_install("requests")
check_and_install("requests_ntlm", "requests_ntlm")
import platform as _platform
if _platform.system() == "Windows":
    check_and_install("pywin32", "win32api")

import openpyxl
import requests
from requests_ntlm import HttpNtlmAuth
requests.packages.urllib3.disable_warnings()

# ── Sabitler ─────────────────────────────────────────────────────────────────
MONTHS_TR = {
    1: "Ocak",    2: "Şubat",  3: "Mart",    4: "Nisan",
    5: "Mayıs",   6: "Haziran",7: "Temmuz",  8: "Ağustos",
    9: "Eylül",  10: "Ekim",  11: "Kasım",  12: "Aralık",
}
CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".poliport_config.json")
APP_VERSION = "1.0"
APP_AUTHOR  = "S. SEYMEN"
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))


# ── Yardımcı fonksiyonlar ─────────────────────────────────────────────────────
def clean_filename_part(s, keep_spaces=False):
    """
    Dosya adı için güvenli hale getir.
    keep_spaces=True  → boşluklar korunur (plaka, sürücü, nakliyeci)
    keep_spaces=False → boşluk kaldırılır (konteyner no — zaten - içeriyor)
    """
    s = str(s or "").strip().upper()
    # Türkçe karakterleri dönüştür
    tr_map = {'Ğ':'G','ğ':'g','Ü':'U','ü':'u','Ş':'S','ş':'s',
              'İ':'I','ı':'i','Ö':'O','ö':'o','Ç':'C','ç':'c'}
    s = re.sub(r'[ĞğÜüŞşİıÖöÇç]', lambda m: tr_map.get(m.group(), ''), s)
    # Windows dosya adında yasak karakterleri kaldır (/ \ * ? : " < > |)
    s = re.sub(r'[\\/*?"<>|]', '', s)
    if not keep_spaces:
        s = re.sub(r'\s+', '', s)   # boşlukları tamamen kaldır (konteyner)
    else:
        s = re.sub(r'\s+', ' ', s).strip()  # çoklu boşluğu tek yap
    return s


def build_filename(plaka, dorse, konteyner, ad, soyad, nakliyeci):
    """
    Dosya adı kuralları (parçalar BOŞLUKLA ayrılır, konteyner kendi içinde - içerir):
      - Konteyner var : CEKICI DORSE KONTEYNER AD SOYAD NAKLIYECI.xlsx
      - Normal TIR   : CEKICI DORSE AD SOYAD NAKLIYECI.xlsx
      - 40 Ayak      : CEKICI AD SOYAD NAKLIYECI.xlsx  (plaka == dorse)
    """
    p = clean_filename_part(plaka,     keep_spaces=False)   # plakada boşluk olmaz
    d = clean_filename_part(dorse,     keep_spaces=False)
    k = clean_filename_part(konteyner, keep_spaces=False)   # TMIU748018-4 → olduğu gibi
    a = clean_filename_part(ad,        keep_spaces=True)    # SAİT → SAIT
    s = clean_filename_part(soyad,     keep_spaces=True)    # EROĞLU → EROGLU
    n = clean_filename_part(nakliyeci, keep_spaces=True)    # EROĞLU NAKLİYAT → EROGLU NAKLIYAT

    parts = [p]
    if d and d != p:
        parts.append(d)
    if k:
        parts.append(k)
    parts.append(a)
    parts.append(s)
    if n:
        parts.append(n)

    return " ".join(parts) + ".xlsx"


def transform_ssrs_url(raw_url):
    """
    SSRS hyperlink URL'ini direkt Excel download URL'ine çevirir.

    Rapordaki URL zaten /ReportServer? formatında geliyor:
      http://servicesrv/ReportServer?%2FGullsEye_Raporlar%2FSEC%2FSecKontrolFormu
        &TransactionID=xxx&rs%3AParameterLanguage=

    Çıktı:
      http://servicesrv/ReportServer?/GullsEye_Raporlar/SEC/SecKontrolFormu
        &TransactionID=xxx&rs:Format=EXCELOPENXML
    """
    if not raw_url:
        return None

    # URL decode et (hem %2F hem %3A gibi encoding'leri çöz)
    url = urllib.parse.unquote(raw_url)

    # Gereksiz SSRS parametrelerini temizle
    url = re.sub(r'[&?]rs:ParameterLanguage=[^&]*', '', url)
    url = re.sub(r'[&?]rs:Format=[^&]*', '', url)
    url = re.sub(r'[&?]rc:Toolbar=[^&]*', '', url)

    # Viewer URL ise direkt URL'e çevir (eski format desteği)
    url = re.sub(
        r'/ReportServer/Pages/ReportViewer\.aspx\?',
        '/ReportServer?',
        url, flags=re.IGNORECASE
    )

    # Sondaki & veya ? temizle
    url = url.rstrip('&').rstrip('?')

    # Excel format parametresi ekle
    sep = '&' if '?' in url else '?'
    return url + sep + 'rs:Format=EXCELOPENXML'


def build_filename(plaka, dorse, konteyner, ad, soyad, nakliyeci):
    """
    Dosya adı kuralları:
      - Konteyner var : CEKICI_DORSE_KONTEYNER_AD_SOYAD_NAKLIYECI.xlsx
      - Normal TIR   : CEKICI_DORSE_AD_SOYAD_NAKLIYECI.xlsx
      - 40 Ayak      : CEKICI_AD_SOYAD_NAKLIYECI.xlsx  (plaka == dorse)
    """
    p = clean_filename_part(plaka)
    d = clean_filename_part(dorse)
    k = clean_filename_part(konteyner)
    a = clean_filename_part(ad)
    s = clean_filename_part(soyad)
    n = clean_filename_part(nakliyeci)

    parts = [p]
    if d and d != p:
        parts.append(d)
    if k:
        parts.append(k)
    parts.append(a)
    parts.append(s)
    if n:
        parts.append(n)

    return "_".join(parts) + ".xlsx"


def build_folder_path(base, tarih):
    """Kontrol Formları / 2026 / Nisan / 29"""
    if not tarih:
        return os.path.join(base, "Kontrol Formları", "Tarihsiz")
    return os.path.join(
        base, "Kontrol Formları",
        str(tarih.year),
        MONTHS_TR[tarih.month],
        str(tarih.day)
    )


# ── Ana uygulama sınıfı ───────────────────────────────────────────────────────
class PoliportApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Poliport Kontrol Formu Arşivi v{APP_VERSION}")
        self.root.geometry("1250x800")
        self.root.configure(bg="#ecf0f1")

        self.records = []
        self.tree_items = []
        self.config = self._load_config()
        self.icon_img = None   # PhotoImage referansı (GC koruması)

        self._load_icon()
        self._setup_styles()
        self._setup_ui()

    # ── İkon yükleme ─────────────────────────────────────────────────────────
    def _load_icon(self):
        icon_path = os.path.join(BASE_DIR, "icon.png")
        ico_path  = os.path.join(BASE_DIR, "icon.ico")
        self._ico_path = None
        try:
            from PIL import Image, ImageTk
            img = Image.open(icon_path)
            # Pencere başlık ikonu
            import platform
            if platform.system() == "Windows":
                if not os.path.exists(ico_path):
                    ico_img = img.resize((256, 256), Image.LANCZOS)
                    ico_img.save(ico_path, format="ICO", sizes=[(256,256),(64,64),(32,32),(16,16)])
                try:
                    self.root.iconbitmap(ico_path)
                    self._ico_path = ico_path
                except Exception:
                    pass
            else:
                # Mac / Linux: wm_iconphoto kullan
                icon_tk = ImageTk.PhotoImage(img.resize((64, 64), Image.LANCZOS))
                self.root.wm_iconphoto(True, icon_tk)
                self._icon_tk = icon_tk  # GC koruması
            # Header'da göstermek için 48x48 tkinter resmi
            thumb = img.resize((48, 48), Image.LANCZOS)
            self.icon_img = ImageTk.PhotoImage(thumb)
        except Exception:
            self.icon_img = None

    # ── Config ───────────────────────────────────────────────────────────────
    def _load_config(self):
        default = {"save_folder": "", "username": "", "password": ""}
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    return {**default, **json.load(f)}
        except Exception:
            pass
        return default

    def _save_config(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # ── Stiller ──────────────────────────────────────────────────────────────
    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("TProgressbar", thickness=12)

    # ── UI ───────────────────────────────────────────────────────────────────
    def _setup_ui(self):
        # ── Header ──
        hdr = tk.Frame(self.root, bg="#1a5276", height=62)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)

        # İkon (varsa)
        if self.icon_img:
            tk.Label(hdr, image=self.icon_img, bg="#1a5276"
                     ).pack(side=tk.LEFT, padx=(10, 4), pady=7)

        tk.Label(hdr, text="Poliport Kontrol Formu Arşivi",
                 font=("Segoe UI", 14, "bold"), fg="white", bg="#1a5276"
                 ).pack(side=tk.LEFT, padx=(4, 0), pady=10)

        tk.Label(hdr, text=f"v{APP_VERSION}", fg="#aed6f1", bg="#1a5276",
                 font=("Segoe UI", 9)).pack(side=tk.RIGHT, padx=14)

        body = tk.Frame(self.root, bg="#ecf0f1")
        body.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

        # ── Ayarlar bölümü ──
        sframe = tk.LabelFrame(body, text="  Dosya Ayarları  ",
                               bg="#ecf0f1", font=("Segoe UI", 9, "bold"),
                               padx=10, pady=8)
        sframe.pack(fill=tk.X, pady=(0, 8))
        sframe.columnconfigure(1, weight=1)

        # Rapor satırı
        tk.Label(sframe, text="Rapor Dosyası:", bg="#ecf0f1",
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=2)
        self.report_var = tk.StringVar()
        tk.Entry(sframe, textvariable=self.report_var, bg="white",
                 font=("Segoe UI", 9)).grid(row=0, column=1, sticky="ew", padx=(6, 4))
        tk.Button(sframe, text="Gözat", command=self.browse_report,
                  width=7).grid(row=0, column=2, padx=2)
        tk.Button(sframe, text="Yükle ▶", command=self.load_report,
                  bg="#2980b9", fg="white", font=("Segoe UI", 9, "bold"),
                  width=8).grid(row=0, column=3, padx=(2, 0))

        # Klasör satırı
        tk.Label(sframe, text="Kayıt Klasörü:", bg="#ecf0f1",
                 font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", pady=2)
        self.folder_var = tk.StringVar(value=self.config.get("save_folder", ""))
        tk.Entry(sframe, textvariable=self.folder_var, bg="white",
                 font=("Segoe UI", 9)).grid(row=1, column=1, sticky="ew", padx=(6, 4))
        tk.Button(sframe, text="Gözat", command=self.browse_folder,
                  width=7).grid(row=1, column=2, padx=2)

        # ── Araç listesi ──
        lframe = tk.LabelFrame(body, text="  Araç Listesi  ",
                               bg="#ecf0f1", font=("Segoe UI", 9, "bold"),
                               padx=8, pady=6)
        lframe.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        # Toolbar
        toolbar = tk.Frame(lframe, bg="#ecf0f1")
        toolbar.pack(fill=tk.X, pady=(0, 4))

        tk.Button(toolbar, text="☑ Tümünü Seç",
                  command=lambda: self.select_all(True)).pack(side=tk.LEFT)
        tk.Button(toolbar, text="☐ Seçimi Kaldır",
                  command=lambda: self.select_all(False)).pack(side=tk.LEFT, padx=4)
        tk.Button(toolbar, text="🔗 Test (1 Araç)",
                  command=self.test_single).pack(side=tk.LEFT, padx=4)

        self.count_lbl = tk.Label(toolbar, text="Rapor bekleniyor…",
                                  fg="#555", bg="#ecf0f1", font=("Segoe UI", 9))
        self.count_lbl.pack(side=tk.LEFT, padx=10)

        self.dl_btn = tk.Button(toolbar, text="⬇  Seçilenleri İndir",
                                command=self.start_download,
                                bg="#1e8449", fg="white",
                                font=("Segoe UI", 10, "bold"), padx=12)
        self.dl_btn.pack(side=tk.RIGHT)

        # Treeview
        cols = ("chk", "tarih", "plaka", "dorse", "konteyner", "tip", "ad", "soyad", "nakliyeci", "durum")
        self.tree = ttk.Treeview(lframe, columns=cols, show="headings",
                                 height=14, selectmode="none")

        hdrs = [("chk","✓",38), ("tarih","Kayıt Tarihi",130),
                ("plaka","Çekici Plaka",105), ("dorse","Dorse Plaka",105),
                ("konteyner","Konteyner No",115),
                ("tip","Araç Tipi",110), ("ad","Sürücü Adı",90),
                ("soyad","Sürücü Soyadı",110), ("nakliyeci","Nakliyeci",130),
                ("durum","Durum",200)]
        for c, h, w in hdrs:
            self.tree.heading(c, text=h)
            anc = "center" if c == "chk" else "w"
            self.tree.column(c, width=w, minwidth=w, anchor=anc)

        self.tree.tag_configure("ok",     background="#d5f5e3")
        self.tree.tag_configure("error",  background="#fadbd8")
        self.tree.tag_configure("nolink", background="#fef9e7", foreground="#7d6608")
        self.tree.tag_configure("active", background="#d6eaf8")

        vsb = ttk.Scrollbar(lframe, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(lframe, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<Button-1>", self.on_tree_click)

        # ── İşlem Durumu ──
        pframe = tk.LabelFrame(body, text="  İşlem Durumu  ",
                               bg="#ecf0f1", font=("Segoe UI", 9, "bold"),
                               padx=8, pady=6)
        pframe.pack(fill=tk.X)

        self.progress = ttk.Progressbar(pframe, mode="determinate")
        self.progress.pack(fill=tk.X, pady=(0, 4))

        log_row = tk.Frame(pframe, bg="#ecf0f1")
        log_row.pack(fill=tk.X)
        self.log_box = tk.Text(log_row, height=5, state=tk.DISABLED,
                               bg="#1e1e1e", fg="#98fb98",
                               font=("Consolas", 9), wrap=tk.WORD)
        lsb = ttk.Scrollbar(log_row, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=lsb.set)
        lsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_box.pack(fill=tk.X, expand=True)

        # ── Footer / İmza ──
        footer = tk.Frame(self.root, bg="#1a5276", height=24)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer.pack_propagate(False)
        tk.Label(footer,
                 text=f"© {datetime.now().year}  {APP_AUTHOR} tarafından hazırlanmıştır",
                 font=("Segoe UI", 8), fg="#aed6f1", bg="#1a5276"
                 ).pack(side=tk.RIGHT, padx=12)

    # ── Dosya/klasör seçimi ───────────────────────────────────────────────────
    def _get_downloads_folder(self):
        """Gerçek İndirilenler/Downloads klasörünü döndürür (Windows/Mac/Linux)."""
        import platform
        home = os.path.expanduser("~")
        if platform.system() == "Windows":
            try:
                import winreg
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                    r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders") as k:
                    path = winreg.QueryValueEx(k, "{374DE290-123F-4565-9164-39C4925E467B}")[0]
                    if os.path.isdir(path):
                        return path
            except Exception:
                pass
            for name in ["Downloads", "İndirilenler"]:
                d = os.path.join(home, name)
                if os.path.isdir(d):
                    return d
        else:
            for name in ["Downloads", "Desktop", "Masaüstü"]:
                d = os.path.join(home, name)
                if os.path.isdir(d):
                    return d
        return home

    def browse_report(self):
        init_dir = self.config.get("last_report_dir", "")
        if not init_dir or not os.path.isdir(init_dir):
            init_dir = self._get_downloads_folder()
        p = filedialog.askopenfilename(
            title="Rapor Dosyasını Seç",
            initialdir=init_dir,
            filetypes=[("Excel Dosyası", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )
        if p:
            self.config["last_report_dir"] = os.path.dirname(p)
            self._save_config()
            self.report_var.set(p)
            self.load_report()

    def browse_folder(self):
        p = filedialog.askdirectory(title="Kontrol Formlarının Kaydedileceği Klasörü Seç")
        if p:
            self.folder_var.set(p)
            self.config["save_folder"] = p
            self._save_config()

    def _ask_credentials(self, retry=False):
        """
        Kimlik bilgisi popup'ı. 401 hatası gelince çağrılır.
        Döndürür: (username, password) ya da (None, None) iptal edilirse.
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("Kimlik Doğrulama Gerekli")
        dialog.geometry("380x200")
        dialog.resizable(False, False)
        dialog.grab_set()
        dialog.configure(bg="#ecf0f1")

        try:
            if hasattr(self, '_ico_path') and self._ico_path:
                dialog.iconbitmap(self._ico_path)
        except Exception:
            pass

        msg = ("Sunucu kimlik doğrulaması istedi.\n"
               "Kullanıcı adı ve şifrenizi girin:")
        if retry:
            msg = "⚠  Kullanıcı adı veya şifre hatalı.\nTekrar deneyin:"

        tk.Label(dialog, text=msg, bg="#ecf0f1", font=("Segoe UI", 9),
                 justify="left").pack(padx=16, pady=(14, 8), anchor="w")

        frame = tk.Frame(dialog, bg="#ecf0f1")
        frame.pack(padx=16, fill=tk.X)

        tk.Label(frame, text="Kullanıcı Adı:", bg="#ecf0f1",
                 font=("Segoe UI", 9), width=14, anchor="w").grid(row=0, column=0, pady=3)
        user_e = tk.Entry(frame, width=28, bg="white", font=("Segoe UI", 9))
        user_e.grid(row=0, column=1)
        saved_user = self.config.get("username", "")
        user_e.insert(0, saved_user)
        # Domain ipucu (Mac'te DOMAIN\kullanici formatı gerekebilir)
        import platform as _pl
        if _pl.system() != "Windows" and "\\" not in saved_user:
            tk.Label(frame, text="Mac: POLISAN\\kullanici", bg="#ecf0f1",
                     font=("Segoe UI", 8), fg="#888").grid(row=0, column=2, padx=4)

        tk.Label(frame, text="Şifre:", bg="#ecf0f1",
                 font=("Segoe UI", 9), width=14, anchor="w").grid(row=1, column=0, pady=3)
        pass_e = tk.Entry(frame, show="●", width=28, bg="white", font=("Segoe UI", 9))
        pass_e.grid(row=1, column=1)
        pass_e.insert(0, self.config.get("password", ""))

        result = [None, None]

        def on_ok():
            result[0] = user_e.get().strip()
            result[1] = pass_e.get()
            # Kaydet
            self.config["username"] = result[0]
            self.config["password"] = result[1]
            self._save_config()
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        btn_frame = tk.Frame(dialog, bg="#ecf0f1")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Bağlan", command=on_ok,
                  bg="#1e8449", fg="white", font=("Segoe UI", 9, "bold"),
                  width=10).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="İptal", command=on_cancel,
                  width=8).pack(side=tk.LEFT)

        user_e.focus()
        pass_e.bind("<Return>", lambda e: on_ok())
        dialog.wait_window()
        return result[0], result[1]

    # ── Log ──────────────────────────────────────────────────────────────────
    def log(self, msg):
        def _do():
            self.log_box.config(state=tk.NORMAL)
            ts = datetime.now().strftime("%H:%M:%S")
            self.log_box.insert(tk.END, f"[{ts}] {msg}\n")
            self.log_box.see(tk.END)
            self.log_box.config(state=tk.DISABLED)
        self.root.after(0, _do)

    # ── Rapor yükleme ─────────────────────────────────────────────────────────
    def load_report(self):
        path = self.report_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Hata", "Geçerli bir Excel rapor dosyası seçin.")
            return

        # Mevcut listeyi temizle
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.records.clear()
        self.tree_items.clear()

        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            # ── Başlık satırını bul ──
            header_row = None
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
                for cell in row:
                    v = str(cell.value or "").strip()
                    if v in ("Kayıt No", "Plaka", "KAYIT NO"):
                        header_row = ri
                        break
                if header_row:
                    break

            if not header_row:
                messagebox.showerror("Hata",
                    "Başlık satırı bulunamadı.\n"
                    "'Kayıt No' veya 'Plaka' sütun adlarını içerdiğinden emin olun.")
                return

            # ── Sütun haritası (whitespace normalize edilmiş) ──
            col_map = {}
            for cell in ws[header_row]:
                if cell.value is not None:
                    # \n ve çoklu boşlukları tek boşluğa indir
                    v = re.sub(r'\s+', ' ', str(cell.value)).strip()
                    if v:
                        col_map[v] = cell.column - 1   # 0 tabanlı

            self.log(f"Başlık satırı: {header_row}")
            self.log(f"Sütunlar: {', '.join(col_map.keys())}")

            # Sütun adı eşleştirme:
            # 1) Tam eşleşme
            # 2) Aday, anahtar içinde geçiyor (cand ⊆ key)
            def fc(*candidates):
                for cand in candidates:
                    cand_n = re.sub(r'\s+', ' ', cand).strip().lower()
                    # Önce tam eşleşme
                    for k, v in col_map.items():
                        if cand_n == k.lower():
                            return v
                    # Sonra: aday tam olarak key içinde geçiyor
                    for k, v in col_map.items():
                        if cand_n in k.lower():
                            return v
                return None

            c_tarih      = fc("Kayıt Tarihi", "Tarih")
            c_departman  = fc("Departman")
            c_plaka      = fc("Plaka")
            c_dorse     = fc("Dorse Plaka", "Dorse")
            c_konteyner = fc("Konteyner No", "Konteyner")
            c_tip       = fc("Araç Tipi", "Araç Tip")
            c_ad        = fc("Sürücü Adı", "Sürücü Ad")
            c_soyad     = fc("Sürücü Soyadı", "Soyadı")
            c_nakliyeci = fc("Nakliyeci")

            # ── Bilinen sabit sütun konumları (Arac_Kayit_Sorgulama formatı) ──
            KNOWN_COLS = {
                "tarih":     1,   # B
                "departman": 5,   # F
                "plaka":     6,   # G
                "dorse":     8,   # I
                "konteyner": 9,   # J
                "tip":       12,  # M
                "ad":        14,  # O
                "soyad":     15,  # P
                "nakliyeci": 16,  # Q
            }
            c_tarih     = c_tarih     if c_tarih     is not None else KNOWN_COLS["tarih"]
            c_departman = c_departman if c_departman is not None else KNOWN_COLS["departman"]
            c_plaka     = c_plaka     if c_plaka     is not None else KNOWN_COLS["plaka"]
            c_dorse     = c_dorse     if c_dorse     is not None else KNOWN_COLS["dorse"]
            c_konteyner = c_konteyner if c_konteyner is not None else KNOWN_COLS["konteyner"]
            c_tip       = c_tip       if c_tip       is not None else KNOWN_COLS["tip"]
            c_ad        = c_ad        if c_ad        is not None else KNOWN_COLS["ad"]
            c_soyad     = c_soyad     if c_soyad     is not None else KNOWN_COLS["soyad"]
            c_nakliyeci = c_nakliyeci if c_nakliyeci is not None else KNOWN_COLS["nakliyeci"]

            self.log(f"Plaka={c_plaka+1}, Dorse={c_dorse+1}, Konteyner={c_konteyner+1}, Nakliyeci={c_nakliyeci+1}")

            # ── Veri satırlarını işle ──
            loaded = link_count = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
                if not any(c.value for c in row):
                    continue   # boş satır atla

                def gv(ci):
                    if ci is not None and ci < len(row):
                        v = row[ci].value
                        return str(v).strip() if v is not None else ""
                    return ""

                # ── Departman filtresi: sadece "Poliport Terminal" ──
                departman = gv(c_departman)
                if "terminal" not in departman.lower():
                    continue

                plaka = gv(c_plaka)
                if not plaka or plaka.lower() == "none":
                    continue

                dorse      = gv(c_dorse)
                konteyner  = gv(c_konteyner)
                tip        = gv(c_tip)
                ad         = gv(c_ad)
                soyad      = gv(c_soyad)
                nakliyeci  = gv(c_nakliyeci)

                # Tarihi al
                tarih_raw = row[c_tarih].value if c_tarih is not None else None
                tarih = None
                if isinstance(tarih_raw, datetime):
                    tarih = tarih_raw
                elif tarih_raw:
                    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M",
                                "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M"):
                        try:
                            tarih = datetime.strptime(str(tarih_raw).strip(), fmt)
                            break
                        except ValueError:
                            continue

                tarih_disp = tarih.strftime("%d.%m.%Y %H:%M") if tarih else str(tarih_raw or "")

                # Hyperlink al
                url = None
                if c_plaka is not None and c_plaka < len(row):
                    hl = row[c_plaka].hyperlink
                    if hl:
                        url = hl.target

                has_link = bool(url)
                if has_link:
                    link_count += 1

                rec = {
                    "tarih": tarih, "tarih_disp": tarih_disp,
                    "plaka": plaka, "dorse": dorse,
                    "konteyner": konteyner,
                    "tip": tip, "ad": ad, "soyad": soyad,
                    "nakliyeci": nakliyeci,
                    "url": url,
                    "selected": has_link,
                    "durum": "⏳ Bekliyor" if has_link else "⚠ Link Yok",
                }
                self.records.append(rec)

                tag = "" if has_link else "nolink"
                iid = self.tree.insert("", tk.END, values=(
                    "☑" if has_link else "–",
                    tarih_disp, plaka, dorse, konteyner,
                    tip, ad, soyad, nakliyeci,
                    rec["durum"],
                ), tags=(tag,))
                self.tree_items.append(iid)
                loaded += 1

            self.count_lbl.config(
                text=f"{loaded} araç yüklendi  |  {link_count} indirilebilir"
            )
            self.log(f"Rapor yüklendi — {loaded} araç, {link_count} linkli.")

        except Exception as e:
            messagebox.showerror("Rapor Yükleme Hatası", str(e))
            self.log(f"HATA: {e}")

    # ── Treeview tıklama ─────────────────────────────────────────────────────
    def on_tree_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item or item not in self.tree_items:
            return
        idx = self.tree_items.index(item)
        rec = self.records[idx]
        if not rec["url"]:
            return
        rec["selected"] = not rec["selected"]
        vals = list(self.tree.item(item, "values"))
        vals[0] = "☑" if rec["selected"] else "☐"
        self.tree.item(item, values=vals)

    def select_all(self, state):
        for i, iid in enumerate(self.tree_items):
            if i < len(self.records) and self.records[i]["url"]:
                self.records[i]["selected"] = state
                vals = list(self.tree.item(iid, "values"))
                vals[0] = "☑" if state else "☐"
                self.tree.item(iid, values=vals)

    # ── Satır durumunu güncelle ───────────────────────────────────────────────
    def _set_row(self, idx, durum, tag=""):
        def _do():
            if idx < len(self.tree_items):
                iid = self.tree_items[idx]
                vals = list(self.tree.item(iid, "values"))
                vals[9] = durum   # durum sütunu: chk(0)..nakliyeci(8) durum(9)
                self.tree.item(iid, values=vals, tags=(tag,) if tag else ())
        self.root.after(0, _do)

    # ── Test (tek araç) ──────────────────────────────────────────────────────
    def test_single(self):
        first = next(
            ((i, r) for i, r in enumerate(self.records) if r.get("url")), None
        )
        if not first:
            messagebox.showwarning("Uyarı", "Test için linkli araç bulunamadı.")
            return
        idx, rec = first
        dl_url = transform_ssrs_url(rec["url"])
        msg = (
            f"Araç: {rec['plaka']}\n\n"
            f"Kaynak URL:\n{rec['url']}\n\n"
            f"Download URL:\n{dl_url}"
        )
        if messagebox.askyesno("URL Testi", msg + "\n\nBağlantıyı test etmek ister misiniz?"):
            threading.Thread(target=self._test_url, args=(idx, rec, dl_url), daemon=True).start()

    def _test_url(self, idx, rec, dl_url):
        session = self._make_session()
        try:
            r = session.head(dl_url, timeout=15, allow_redirects=True)
            if r.status_code == 200:
                ct = r.headers.get("Content-Type", "")
                self.root.after(0, lambda: messagebox.showinfo(
                    "Test Başarılı", f"✓ Bağlantı başarılı!\nHTTP 200\nContent-Type: {ct}"))
            else:
                self.root.after(0, lambda: messagebox.showerror(
                    "Test Başarısız", f"HTTP {r.status_code}\nURL: {dl_url}"))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Bağlantı Hatası", str(e)))

    # ── Session oluştur ──────────────────────────────────────────────────────
    def _make_session(self, username=None, password=None):
        """
        Önce Windows oturumu (SSPI/NTLM) ile dener.
        Kayıtlı kimlik varsa ya da parametre geçildiyse onu kullanır.
        """
        session = requests.Session()
        session.verify = False

        u = username or self.config.get("username", "")
        p = password or self.config.get("password", "")

        if u and p:
            session.auth = HttpNtlmAuth(u, p)
        else:
            import platform
            if platform.system() == "Windows":
                # Windows integrated auth (SSPI) — domain PC'lerde şifresiz çalışır
                try:
                    from requests_negotiate_sspi import HttpNegotiateAuth
                    session.auth = HttpNegotiateAuth()
                except ImportError:
                    pass   # Auth olmadan dene; 401 gelirse popup açılır
            # Mac/Linux: auth olmadan dene; 401 gelirse popup açılır
        return session

    # ── İndirme ──────────────────────────────────────────────────────────────
    def start_download(self):
        save_folder = self.folder_var.get().strip()
        if not save_folder:
            messagebox.showerror("Hata", "Lütfen kayıt klasörü seçin.")
            return
        selected = [(i, r) for i, r in enumerate(self.records)
                    if r.get("selected") and r.get("url")]
        if not selected:
            messagebox.showwarning("Uyarı", "İndirilecek araç seçilmedi.")
            return
        self.dl_btn.config(state=tk.DISABLED, text="⏳ İndiriliyor…")
        self.progress.config(maximum=len(selected), value=0)
        threading.Thread(
            target=self._download_thread,
            args=(selected, save_folder),
            daemon=True
        ).start()

    def _download_thread(self, selected, save_folder):
        session = self._make_session()
        ok = fail = 0
        auth_confirmed = False   # credentials başarılı doğrulandı mı

        for prog, (rec_idx, rec) in enumerate(selected):
            plaka = rec["plaka"]
            self.log(f"→ {plaka}…")
            self._set_row(rec_idx, "⏬ İndiriliyor…", "active")

            dl_url = transform_ssrs_url(rec["url"])
            if not dl_url:
                self.log(f"  ✗ URL dönüştürülemedi: {plaka}")
                self._set_row(rec_idx, "✗ URL Hatası", "error")
                fail += 1
                continue

            try:
                resp = session.get(dl_url, timeout=45, stream=True)

                if resp.status_code == 200:
                    folder = build_folder_path(save_folder, rec["tarih"])
                    os.makedirs(folder, exist_ok=True)

                    fname = build_filename(
                        rec["plaka"], rec["dorse"], rec["konteyner"],
                        rec["ad"], rec["soyad"], rec["nakliyeci"]
                    )
                    fpath = os.path.join(folder, fname)

                    # Aynı isimli dosya varsa sona numara ekle
                    base, ext = os.path.splitext(fpath)
                    counter = 1
                    while os.path.exists(fpath):
                        fpath = f"{base}_{counter}{ext}"
                        counter += 1

                    with open(fpath, "wb") as f:
                        for chunk in resp.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)

                    self.log(f"  ✓ {fname}")
                    self._set_row(rec_idx, f"✓ {fname}", "ok")
                    ok += 1

                elif resp.status_code == 401:
                    # 401: popup göster, max 3 deneme, başarılı olunca auth_confirmed=True
                    if not auth_confirmed:
                        self.log(f"  ⚠ Yetki hatası — kimlik bilgisi isteniyor...")
                        import time
                        saved = False
                        for attempt in range(3):
                            cred_result = [None, None]
                            done_flag   = [False]

                            def ask_cred(retry=attempt > 0):
                                u, p = self._ask_credentials(retry=retry)
                                cred_result[0] = u if u else ""
                                cred_result[1] = p if p else ""
                                done_flag[0] = True

                            self.root.after(0, ask_cred)
                            for _ in range(120):
                                time.sleep(0.5)
                                if done_flag[0]:
                                    break

                            if not cred_result[0]:  # iptal
                                self.log(f"  ✗ İptal edildi.")
                                self._set_row(rec_idx, "✗ İptal", "error")
                                fail += 1
                                break

                            session = self._make_session(cred_result[0], cred_result[1])
                            try:
                                resp2 = session.get(dl_url, timeout=45, stream=True)
                            except Exception as e:
                                self.log(f"  ✗ {e}")
                                self._set_row(rec_idx, "✗ Hata", "error")
                                fail += 1
                                break

                            if resp2.status_code == 200:
                                folder = build_folder_path(save_folder, rec["tarih"])
                                os.makedirs(folder, exist_ok=True)
                                fname = build_filename(rec["plaka"], rec["dorse"],
                                                      rec["konteyner"], rec["ad"],
                                                      rec["soyad"], rec["nakliyeci"])
                                fpath = os.path.join(folder, fname)
                                base, ext = os.path.splitext(fpath)
                                counter = 1
                                while os.path.exists(fpath):
                                    fpath = f"{base}_{counter}{ext}"
                                    counter += 1
                                with open(fpath, "wb") as f:
                                    for chunk in resp2.iter_content(8192):
                                        if chunk: f.write(chunk)
                                self.log(f"  ✓ {fname}")
                                self._set_row(rec_idx, f"✓ {fname}", "ok")
                                ok += 1
                                auth_confirmed = True  # kimlik doğrulandı
                                saved = True
                                break
                            else:
                                self.log(f"  ✗ Kimlik hatalı (deneme {attempt+1}/3)")
                                if attempt == 2:
                                    self.log("  ✗ 3 deneme başarısız, iptal ediliyor.")
                                    self._set_row(rec_idx, "✗ Kimlik Hatası", "error")
                                    fail += 1
                    else:
                        self.log(f"  ✗ Yetki hatası (401): {plaka}")
                        self._set_row(rec_idx, "✗ Yetki Hatası", "error")
                        fail += 1

                else:
                    self.log(f"  ✗ HTTP {resp.status_code}: {plaka}")
                    self._set_row(rec_idx, f"✗ HTTP {resp.status_code}", "error")
                    fail += 1

            except requests.exceptions.ConnectionError:
                self.log(f"  ✗ Bağlantı hatası — sunucu erişilemiyor: {plaka}")
                self._set_row(rec_idx, "✗ Bağlantı Hatası", "error")
                fail += 1

            except Exception as e:
                self.log(f"  ✗ {plaka}: {e}")
                self._set_row(rec_idx, "✗ Hata", "error")
                fail += 1

            self.root.after(0, lambda v=prog + 1: self.progress.config(value=v))

        def _done():
            self.dl_btn.config(state=tk.NORMAL, text="⬇  Seçilenleri İndir")
            self.log(f"=== Tamamlandı: {ok} başarılı, {fail} hatalı ===")
            messagebox.showinfo(
                "İndirme Tamamlandı",
                f"✓ {ok} form başarıyla indirildi\n✗ {fail} hata\n\n"
                f"Klasör: {save_folder}"
            )
        self.root.after(0, _done)


# ── Başlat ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()

    # Windows DPI düzeltmesi
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = PoliportApp(root)
    root.mainloop()
